'''
Created on May 8, 2020

@author: NOORSHAVALI

1) page timeouts
2) Screenshots save 
3) xls file reading for data driven testing
'''
import shutil  
import openpyxl as xllib
import logging,datetime,re,os,shutil
import sys

class UtilCalss(object):
    '''
    classdocs
    '''
    PAGE_LOAD_TIMEOUT_DEFAULT = 20
    PAGE_LOAD_TIMEOUT_QUICK = 3  #sec
    PAGE_LOAD_TIMEOUT_MEDIUM = 5
    PAGE_LOAD_TIMEOUT_LARGE = 9
    
    IMPLICIT_WAIT = 20
    THREAD_WAIT_TIME = 10
    FLUENT_WAIT_TIME = 10  # explicit wait


    def __init__(self):
        '''
        Constructor
        '''
    
    def takescreenshot(self,driver,methodname):  
        pass
    @staticmethod
    def takescreenshot_failure(driver,methodname):
        timestamp = re.sub('[-:. ]','',str(datetime.datetime.now()))[:-3]
        failedtestscreen = os.getcwd() + methodname + timestamp + ".png"
        distdir = 'C:/Users/NOORSHAVALI/eclipse-workspace/HelloPythonWorld/output'
        driver.get_screenshot_as_file(failedtestscreen)
        shutil.copy(failedtestscreen,distdir)  
            
        
    @staticmethod
    def read_data_from_excel(sheetname = None):
        try:
            workbookname = xllib.load_workbook("C:\Users\NOORSHAVALI\Desktop\DynamicRenewel11.xlsx")
            sheetname = workbookname.sheetnames[-1].encode()
            sheetobj = workbookname.get_sheet_by_name(sheetname)
            rows,columns = sheetobj.max_row,sheetobj.max_column
            
            #" wb.get_sheet_names()"
            for row in range(2,rows+1):
                lst = []
                for col in range(1,columns+1):
                    val = sheetobj.cell(row=row,column=col).value
                    if isinstance(val,basestring):
                        lst.append(val.encode())
                    else:
                        lst.append(int(val))
                UtilCalss.getlogger(UtilCalss()).debug("DDT-excel-sheet record : {}".format(lst))
                yield lst
        except IOError as e:
            UtilCalss.getlogger(UtilCalss()).debug("Exception : {}".format(e.args))
        except Exception as e:
            UtilCalss.getlogger(UtilCalss()).debug("Exception : {}".format(e.args))
        else:
            workbookname.close()
            UtilCalss.getlogger(UtilCalss()).debug("Closing sheet")
        finally:
            UtilCalss.getlogger(UtilCalss()).debug("DDT is completed using excel-sheet")
            
    @staticmethod
    def util_excel(sheetname = None):
        try:
            if sheetname == None:
                workbookname = xllib.load_workbook("../../POM_webdriver_bdd/POM_webdriver_bdd_data/DynamicRenewel11.xlsx")
                sheetname = workbookname.sheetnames[-1].encode()
                sheetobj = workbookname.get_sheet_by_name(sheetname)
            else:
                sheetobj = sheetname
            
            numrows,numcolumns = sheetobj.max_row,sheetobj.max_column
            for row in sheetobj.iter_rows(min_row=2,min_col=1,max_row=numrows,max_col=numcolumns):
                lst = []
                for cell in row:
                    val = cell.value
                    if isinstance(val,basestring):
                        lst.append(val.encode())
                    else:
                        lst.append(int(val))
                yield lst
        except IOError as e:
            UtilCalss.getlogger(UtilCalss()).debug("Exception : {}".format(e.args))
        except Exception as e:
            UtilCalss.getlogger(UtilCalss()).debug("Exception : {} - {} ".format(e.args,sys.exc_info()))
        else:
            if sheetname == None:
                workbookname.close()
            UtilCalss.getlogger(UtilCalss()).debug("Closing sheet")
        finally:
            UtilCalss.getlogger(UtilCalss()).debug("DDT is completed using excel-sheet")    
                    
    pass    
    
    @staticmethod
    def util_excel_sheetobj(sheetname = None):
        try:
            workbookname = xllib.load_workbook("../../POM_webdriver_bdd/POM_webdriver_bdd_data/DynamicRenewel11.xlsx")
            sheetname = workbookname.sheetnames[-1].encode()
            if sheetname is not None:
                sheetobj = workbookname.get_sheet_by_name(sheetname)
            else:
                sheetobj = workbookname.get_sheet_by_name("Sheet6")
        except Exception as e:
            UtilCalss.getlogger(UtilCalss()).debug("Exception : {} - {} ".format(e.args,sys.exc_info()))
   
        else:
            return sheetobj
        
    """            
    lst1 = []
        for col in range(1,cols+1):
            val = sheetname.cell(row=row,column=col).value
            if isinstance(val,basestring):
                lst.append(val.encode())
            else:
                lst.append(int(val))

        yield lst

    """               
        
    @staticmethod
    def  getlogger(self):
        logger = logging.getLogger(__name__)
        #logger.setLevel(logging.INFO)
        logger.setLevel(logging.DEBUG)

        formatter = logging.Formatter('%(asctime)s:%(levelname)s:%(name)s:%(message)s')

        file_handler = logging.FileHandler('../../POM_webdriver_bdd/POM_webdriver_bdd_logs/employee.log')
        file_handler.setFormatter(formatter)

        logger.addHandler(file_handler)
        return logger
        

'''
C:\Users\NOORSHAVALI\eclipse-workspace\HelloPythonWorld\POM_webdriver_bdd_logs
traceback to get function name        
        
from functools import import wraps

def tmp_wrap(func):
@wraps(func)
def tmp(*args,**kwargs):
print func.__name__
return func(*args,**Kwargs)
return tmp

@tmp_wrap
def my_funcky_name():
print "STUB"

my_funky_name()



'''